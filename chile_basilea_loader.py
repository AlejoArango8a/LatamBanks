#!/usr/bin/env python3
"""
Chile CMF — Adecuación de Capital (Basilea III) xlsx → CockroachDB.

Fuente: https://www.cmfchile.cl/portal/estadisticas/626/w4-propertyvalue-43980.html
Workbook mensual por banco: INDICADORES + CAPITAL REGULATORIO Y ACTIVOS.

Almacenamiento (HANDOFF_CL_CMF_MaxData_Blueprint.md §7):
  tipo='q1'  → ratios % ×100 (convención UY/US)
  tipo='x1'  → stocks en pesos (millones CMF × 1e6)

Cuentas: CL_B3_* (ver PLAN_LABELS).

Uso:
  python chile_basilea_loader.py                  # incremental (listing − ya cargados)
  python chile_basilea_loader.py --all --force
  python chile_basilea_loader.py --from 202201 --to 202605
  python chile_basilea_loader.py --article-id 112239
  python chile_basilea_loader.py --xlsx-path /tmp/file.xlsx
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import re
import unicodedata
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import load_workbook

from chile_loader import (
    MONTH_TO_MM,
    _http_get,
    parse_period_from_cmf_title,
)
from cmf_loader import get_connection

load_dotenv(Path(__file__).parent / ".env")

log = logging.getLogger("chile_basilea")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

COUNTRY = "CL"
LISTING_BASILEA_URL = (
    "https://www.cmfchile.cl/portal/estadisticas/626/w4-propertyvalue-43980.html"
)
MILLIONS_SCALE = 1_000_000
BATCH = 500
SISTEMA_COD = 999

# Ratios → tipo q1 (percent × 100)
Q1_ACCOUNTS = {
    "CL_B3_PE_APR": "Patrimonio Efectivo / APR",
    "CL_B3_T1_APR": "Capital Nivel 1 / APR",
    "CL_B3_CET1_APR": "Capital Básico (CET1) / APR",
    "CL_B3_LEV": "Capital Básico / Activos Totales Regulatorios",
    "CL_B3_BUF_DEF": "Déficit colchones conservación + contra-cíclico",
    "CL_B3_CLASS": "Clasificación de solvencia (art. 61 LGB)",
}

# Stocks → tipo x1 (pesos)
X1_ACCOUNTS = {
    "CL_B3_CET1": "Capital Básico (CET1)",
    "CL_B3_AT1": "Capital Adicional Nivel 1",
    "CL_B3_T1": "Capital Nivel 1",
    "CL_B3_T2": "Capital Nivel 2",
    "CL_B3_PE": "Patrimonio Efectivo",
    "CL_B3_ATR": "Activos Totales Regulatorios",
    "CL_B3_APR": "Activos Ponderados por Riesgo (APR)",
    "CL_B3_APRC": "APR Crédito",
    "CL_B3_APRM": "APR Mercado",
    "CL_B3_APRO": "APR Operacional",
}

PLAN_LABELS = {**Q1_ACCOUNTS, **X1_ACCOUNTS}

# Hard aliases when fuzzy match against instituciones fails
NAME_ALIASES: dict[str, int] = {
    "banco bice": 28,
    "banco btg pactual chile": 59,
    "banco consorcio": 55,
    "banco de chile": 1,
    "banco de credito e inversiones": 16,
    "banco de crédito e inversiones": 16,
    "banco del estado de chile": 12,
    "banco falabella": 51,
    "banco internacional": 9,
    "banco itau chile": 39,
    "banco itaú chile": 39,
    "banco ripley": 53,
    "banco santander chile": 37,
    "banco santander-chile": 37,
    "bank of china agencia en chile": 61,
    "bank of china, agencia en chile": 61,
    "china construction bank": 60,
    "china construction bank agencia en chile": 60,
    "china construction bank, agencia en chile": 60,
    "hsbc bank chile": 31,
    "hsbc bank (chile)": 31,
    "jp morgan chase bank n a": 41,
    "jp morgan chase bank, n a": 41,
    "jp morgan chase bank, n.a.": 41,
    "jp morgan chase bank n.a.": 41,
    "scotiabank chile": 14,
    "tanner banco digital": 62,
    "banco security": 49,
    "itau corpbanca": 39,
    "itaú corpbanca": 39,
    "sistema bancario": SISTEMA_COD,
    "total sistema financiero": SISTEMA_COD,
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _header_norm(s: str) -> str:
    """Normalize header text; keep 'Nivel 1' digits (footnotes already stripped by _norm)."""
    return _norm(s)


def parse_basilea_listing(html: str) -> dict[int, str]:
    """aid → YYYYMM from propertyvalue-43980 cards."""
    out: dict[int, str] = {}
    if not html:
        return out
    for m in re.finditer(r"\baid-(\d+)\b", html):
        aid = int(m.group(1))
        window = html[m.start() : m.start() + 1200]
        text = re.sub(r"<[^>]+>", " ", window)
        text = re.sub(r"\s+", " ", text)
        per = parse_period_from_cmf_title(text)
        if not per:
            # Titles are often just "Adecuación … Capital Mayo 2026"
            m2 = re.search(
                r"\b([A-Za-zÁÉÍÓÚáéíóúüÜ]+)\s+(\d{4})\b",
                text,
            )
            if m2:
                mm = MONTH_TO_MM.get(m2.group(1).lower())
                if mm:
                    per = f"{m2.group(2)}{mm}"
        if per:
            out.setdefault(aid, per)
    return out


def scrape_basilea_listing(url: str = LISTING_BASILEA_URL) -> dict[int, str]:
    try:
        html = _http_get(url, timeout=120).decode("utf-8", "ignore")
    except Exception as e:
        log.warning("Basilea listing scrape failed: %s", e)
        return {}
    found = parse_basilea_listing(html)
    log.info(
        "Basilea listing → %d aid→period (newest %s)",
        len(found),
        ", ".join(f"{a}→{p}" for a, p in sorted(found.items(), key=lambda kv: kv[1], reverse=True)[:5]),
    )
    return found


def download_article_xlsx(article_id: int) -> bytes | None:
    for portal in (626, 617):
        url = (
            f"https://www.cmfchile.cl/portal/estadisticas/{portal}/"
            f"articles-{article_id}_recurso_1.xlsx"
        )
        try:
            data = _http_get(url, timeout=120)
        except Exception:
            continue
        if data[:2] == b"PK":
            return data
    return None


def period_from_workbook(wb) -> str | None:
    for ws in wb.worksheets:
        for r in range(1, 12):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if not v or not isinstance(v, str):
                    continue
                per = parse_period_from_cmf_title(v)
                if per:
                    return per
                m = re.search(
                    r"(?:MES\s+(?:DE\s+)?)?([A-Za-zÁÉÍÓÚáéíóúüÜ]+)\s+(?:DE\s+)?(\d{4})",
                    v,
                    flags=re.IGNORECASE,
                )
                if m:
                    mm = MONTH_TO_MM.get(m.group(1).lower())
                    if mm:
                        return f"{m.group(2)}{mm}"
    return None


def _find_indicadores_sheet(wb):
    for ws in wb.worksheets:
        if "indicadores" in _norm(ws.title):
            return ws
    return wb.worksheets[0]


def _find_capital_sheet(wb):
    """Prefer 'CAPITAL REGULATORIO Y ACTIVOS'; never pick LÍMITES."""
    for ws in wb.worksheets:
        n = _norm(ws.title)
        if "capital regulatorio" in n and "activos" in n:
            return ws
    for ws in wb.worksheets:
        n = _norm(ws.title)
        if "capital regulatorio" in n and "limites" not in n and "limite" not in n:
            return ws
    return None


def _collect_col_headers(ws, max_row: int = 16, max_col: int | None = None) -> dict[int, list[str]]:
    """col → list of non-empty header strings from top rows (stop before bank names)."""
    mc = max_col or min(ws.max_column or 40, 60)
    out: dict[int, list[str]] = {}
    bank_re = re.compile(
        r"^(banco|bank|scotiabank|hsbc|jp\s*morgan|china|tanner|sistema|ita[uú])",
        re.I,
    )
    for r in range(1, max_row + 1):
        # Bank names may sit in col B or C depending on workbook vintage.
        stop = False
        for nc in (2, 3):
            name_cell = ws.cell(r, nc).value
            if isinstance(name_cell, str) and bank_re.match(name_cell.strip()):
                stop = True
                break
        if stop:
            break
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                continue
            s = str(v).strip()
            if not s or s == ":":
                continue
            if len(s) > 120:
                continue
            out.setdefault(c, []).append(s)
    return out


def _pick_col(headers: dict[int, list[str]], predicates: list) -> int | None:
    """
    predicates: list of callables(norm_header) -> bool, tried in order.
    Prefer the rightmost matching column (totals often sit to the right of components).
    """
    for pred in predicates:
        hits = []
        for col, labels in headers.items():
            for lab in labels:
                if pred(_header_norm(lab)):
                    hits.append(col)
                    break
        if hits:
            return max(hits)
    return None


def _map_indicator_cols(headers: dict[int, list[str]]) -> dict[str, int]:
    m: dict[str, int] = {}
    pe = _pick_col(
        headers,
        [lambda h: "patrimonio efectivo" in h and "ponderados" in h],
    )
    t1 = _pick_col(
        headers,
        [lambda h: "capital nivel 1" in h and "ponderados" in h and "basico" not in h],
    )
    cet1 = _pick_col(
        headers,
        [lambda h: "capital basico" in h and "ponderados" in h],
    )
    lev = _pick_col(
        headers,
        [
            lambda h: "capital basico" in h and ("totales regulatorios" in h or "activos totales" in h),
        ],
    )
    buf = _pick_col(
        headers,
        [lambda h: "deficit" in h and "colchon" in h],
    )
    klass = _pick_col(
        headers,
        [lambda h: "clasificacion" in h and "solvencia" in h],
    )
    if pe:
        m["CL_B3_PE_APR"] = pe
    if t1:
        m["CL_B3_T1_APR"] = t1
    if cet1:
        m["CL_B3_CET1_APR"] = cet1
    if lev:
        m["CL_B3_LEV"] = lev
    if buf:
        m["CL_B3_BUF_DEF"] = buf
    if klass:
        m["CL_B3_CLASS"] = klass
    return m


def _map_capital_cols(headers: dict[int, list[str]]) -> dict[str, int]:
    m: dict[str, int] = {}

    def is_cet1_total(h: str) -> bool:
        # Prefer "Capital Básico (7)" leaf; exclude ratios and " + "
        if "ponderados" in h or "activos" in h:
            return False
        if "+" in h:
            return False
        return bool(re.fullmatch(r"capital basico( \d+)?", h))

    def is_at1(h: str) -> bool:
        if "ponderados" in h or "+" in h:
            return False
        return h == "capital adicional nivel 1" or re.fullmatch(r"capital adicional nivel 1( \d+)?", h) is not None

    def is_t1(h: str) -> bool:
        if "ponderados" in h:
            return False
        if "capital basico" in h and "capital adicional" in h:
            return True
        return bool(re.fullmatch(r"capital nivel 1( \d+)?", h))

    def is_t2(h: str) -> bool:
        if "ponderados" in h or "bonos" in h or "provisiones" in h:
            return False
        return bool(re.fullmatch(r"capital nivel 2( \d+)?", h))

    def is_pe(h: str) -> bool:
        if "ponderados" in h:
            return False
        if "patrimonio efectivo" in h and "activos" not in h:
            return True
        return "capital nivel 1" in h and "capital nivel 2" in h and "+" not in h  # + already spaced

    cet1 = _pick_col(headers, [is_cet1_total])
    at1 = _pick_col(headers, [is_at1])
    t1 = _pick_col(headers, [is_t1])
    t2 = _pick_col(headers, [is_t2])
    pe = _pick_col(headers, [is_pe])
    atr = _pick_col(headers, [lambda h: "activos totales regulatorios" in h])
    aprc = _pick_col(headers, [lambda h: "ponderados por riesgo de credito" in h])
    aprm = _pick_col(headers, [lambda h: "ponderados por riesgo de mercado" in h])
    apro = _pick_col(headers, [lambda h: "ponderados por riesgo operacional" in h])
    def is_apr_total(h: str) -> bool:
        if not h.startswith("activos ponderados por riesgo"):
            return False
        if any(x in h for x in ("credito", "mercado", "operacional")):
            return False
        return True

    apr = _pick_col(headers, [is_apr_total])
    if cet1:
        m["CL_B3_CET1"] = cet1
    if at1:
        m["CL_B3_AT1"] = at1
    if t1:
        m["CL_B3_T1"] = t1
    if t2:
        m["CL_B3_T2"] = t2
    if pe:
        m["CL_B3_PE"] = pe
    if atr:
        m["CL_B3_ATR"] = atr
    if aprc:
        m["CL_B3_APRC"] = aprc
    if aprm:
        m["CL_B3_APRM"] = aprm
    if apro:
        m["CL_B3_APRO"] = apro
    if apr:
        m["CL_B3_APR"] = apr
    return m


_BANK_NAME_RE = re.compile(
    r"^(banco\b|bank\b|scotiabank\b|hsbc\b|jp\s*morgan\b|china\b|tanner\b|sistema\b|ita[uú]\b)",
    re.I,
)


def _detect_name_col(ws) -> int:
    """CMF workbooks put institution names in col B or C depending on vintage."""
    for nc in (2, 3):
        hits = 0
        for r in range(1, min(40, (ws.max_row or 0) + 1)):
            v = ws.cell(r, nc).value
            if isinstance(v, str) and _BANK_NAME_RE.match(v.strip()):
                hits += 1
                if hits >= 2:
                    return nc
    return 2


def _iter_bank_rows(ws, name_col: int | None = None):
    nc = name_col or _detect_name_col(ws)
    for r in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(r, nc).value
        if not v or not isinstance(v, str):
            continue
        name = v.strip()
        if not name or len(name) > 90 or name.startswith("("):
            continue
        if name.lower().startswith("nota"):
            continue
        if not _BANK_NAME_RE.match(name):
            continue
        yield r, name


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("%", "").replace(",", ".")
        if not s:
            return None
        # Solvency class letters (art. 61 LGB) → 1/2/3
        letter = s.upper()
        if letter in {"A", "B", "C"}:
            return float({"A": 1, "B": 2, "C": 3}[letter])
        try:
            return float(s)
        except ValueError:
            return None
    return None


def build_name_index(conn) -> dict[str, int]:
    cur = conn.cursor()
    cur.execute(
        "SELECT codigo, razon_social FROM instituciones WHERE country = %s",
        (COUNTRY,),
    )
    idx: dict[str, int] = dict(NAME_ALIASES)
    for codigo, razon in cur.fetchall():
        n = _norm(razon)
        if n:
            idx.setdefault(n, int(codigo))
        # strip trailing footnotes already handled by _norm
    return idx


def resolve_ins_cod(name: str, index: dict[str, int]) -> int | None:
    n = _norm(name)
    if n in index:
        return index[n]
    # soft contains
    for key, cod in index.items():
        if key and (key in n or n in key):
            return cod
    return None


def parse_basilea_xlsx(xlsx_bytes: bytes, name_index: dict[str, int] | None = None):
    """
    Returns (periodo, rows, unmatched_names).
    rows: list of (periodo, tipo, ins_cod, cuenta, monto_total)
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    periodo = period_from_workbook(wb)
    if not periodo:
        raise ValueError("Could not detect period from Basilea workbook title")

    name_index = name_index or dict(NAME_ALIASES)
    by_bank: dict[int, dict[str, float]] = {}
    unmatched: list[str] = []

    ind = _find_indicadores_sheet(wb)
    ind_map = _map_indicator_cols(_collect_col_headers(ind))
    if "CL_B3_CET1_APR" not in ind_map and "CL_B3_PE_APR" not in ind_map:
        raise ValueError(f"INDICADORES columns not found in sheet {ind.title!r}")

    for r, name in _iter_bank_rows(ind):
        cod = resolve_ins_cod(name, name_index)
        if cod is None:
            unmatched.append(name)
            continue
        bucket = by_bank.setdefault(cod, {})
        for acct, col in ind_map.items():
            val = _num(ind.cell(r, col).value)
            if val is None:
                continue
            bucket[acct] = val

    cap = _find_capital_sheet(wb)
    if cap is not None:
        cap_map = _map_capital_cols(_collect_col_headers(cap))
        for r, name in _iter_bank_rows(cap):
            cod = resolve_ins_cod(name, name_index)
            if cod is None:
                if name not in unmatched:
                    unmatched.append(name)
                continue
            bucket = by_bank.setdefault(cod, {})
            for acct, col in cap_map.items():
                val = _num(cap.cell(r, col).value)
                if val is None:
                    continue
                bucket[acct] = val
            # Prefer APR = APRC+APRM+APRO when components exist (layout varies).
            parts = [
                bucket.get("CL_B3_APRC"),
                bucket.get("CL_B3_APRM"),
                bucket.get("CL_B3_APRO"),
            ]
            if any(p is not None for p in parts):
                bucket["CL_B3_APR"] = sum(p or 0 for p in parts)

    rows = []
    for cod, vals in by_bank.items():
        for acct, raw in vals.items():
            if acct in Q1_ACCOUNTS:
                # percent × 100; classification is already an integer grade
                if acct == "CL_B3_CLASS":
                    monto = int(round(raw))
                else:
                    monto = int(round(raw * 100))
                rows.append((periodo, "q1", cod, acct, monto))
            elif acct in X1_ACCOUNTS:
                monto = int(round(raw * MILLIONS_SCALE))
                rows.append((periodo, "x1", cod, acct, monto))
    return periodo, rows, unmatched


def wipe_basilea_period(conn, periodo: str):
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM datos_financieros WHERE country = %s AND periodo = %s "
        "AND cuenta LIKE 'CL_B3_%%'",
        (COUNTRY, periodo),
    )
    conn.commit()
    return cur.rowcount


def upsert_plan(conn):
    cur = conn.cursor()
    rows = [(COUNTRY, cuenta, label) for cuenta, label in PLAN_LABELS.items()]
    psycopg2.extras.execute_values(
        cur,
        "INSERT INTO plan_cuentas (country, cuenta, descripcion) VALUES %s "
        "ON CONFLICT (country, cuenta) DO UPDATE SET descripcion = EXCLUDED.descripcion",
        rows,
    )
    conn.commit()


def insert_rows(conn, rows: list[tuple]):
    if not rows:
        return 0
    sql = (
        "INSERT INTO datos_financieros "
        "(country, periodo, tipo, ins_cod, cuenta, monto_clp, monto_uf, monto_tc, monto_ext, monto_total) "
        "VALUES %s "
        "ON CONFLICT (country, periodo, tipo, ins_cod, cuenta) DO UPDATE SET "
        "monto_total = EXCLUDED.monto_total"
    )
    tuples = [
        (COUNTRY, per, tipo, ins, cuenta, 0, 0, 0, 0, monto)
        for per, tipo, ins, cuenta, monto in rows
    ]
    cur = conn.cursor()
    for i in range(0, len(tuples), BATCH):
        psycopg2.extras.execute_values(cur, sql, tuples[i : i + BATCH])
    conn.commit()
    return len(tuples)


def basilea_periods_loaded(conn) -> set[str]:
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT periodo FROM datos_financieros "
        "WHERE country = %s AND cuenta = 'CL_B3_CET1_APR'",
        (COUNTRY,),
    )
    return {r[0] for r in cur.fetchall()}


def load_xlsx_bytes(
    xlsx_bytes: bytes,
    *,
    force: bool = False,
    dry_run: bool = False,
    expected_period: str | None = None,
) -> str | None:
    conn = get_connection()
    try:
        name_index = build_name_index(conn)
        periodo, rows, unmatched = parse_basilea_xlsx(xlsx_bytes, name_index)
        if expected_period and expected_period != periodo:
            log.warning(
                "Period mismatch listing=%s workbook=%s — using workbook",
                expected_period,
                periodo,
            )
        loaded = basilea_periods_loaded(conn)
        if periodo in loaded and not force:
            log.info("Skip Basilea %s — already loaded (use --force)", periodo)
            return None
        if unmatched:
            log.warning("Unmatched bank names (%d): %s", len(unmatched), unmatched[:8])
        q1 = sum(1 for r in rows if r[1] == "q1")
        x1 = sum(1 for r in rows if r[1] == "x1")
        banks = len({r[2] for r in rows})
        log.info(
            "Basilea %s · banks=%d · rows=%d (q1=%d x1=%d)%s",
            periodo,
            banks,
            len(rows),
            q1,
            x1,
            " [dry-run]" if dry_run else "",
        )
        if dry_run:
            return periodo
        upsert_plan(conn)
        if force or periodo in loaded:
            n = wipe_basilea_period(conn, periodo)
            log.info("Wiped %d prior CL_B3_* rows for %s", n, periodo)
        insert_rows(conn, rows)
        return periodo
    finally:
        conn.close()


def run_basilea_incremental(*, force: bool = False, dry_run: bool = False) -> list[str]:
    """Download missing months from the CMF Basilea listing."""
    listing = scrape_basilea_listing()
    if not listing:
        log.warning("No Basilea articles discovered")
        return []
    conn = get_connection()
    try:
        have = basilea_periods_loaded(conn) if not force else set()
    finally:
        conn.close()
    # Prefer highest AID per period
    by_period: dict[str, int] = {}
    for aid, per in listing.items():
        prev = by_period.get(per)
        if prev is None or aid > prev:
            by_period[per] = aid
    todo = sorted(
        ((per, aid) for per, aid in by_period.items() if force or per not in have),
        key=lambda t: t[0],
    )
    log.info("Basilea incremental · to_load=%d (force=%s)", len(todo), force)
    done = []
    for per, aid in todo:
        data = download_article_xlsx(aid)
        if not data:
            log.warning("Download failed aid=%s period=%s", aid, per)
            continue
        try:
            got = load_xlsx_bytes(
                data, force=force, dry_run=dry_run, expected_period=per
            )
        except Exception as e:
            log.exception("Parse/load failed aid=%s: %s", aid, e)
            continue
        if got:
            done.append(got)
    log.info("Basilea incremental finished · loaded=%d", len(done))
    return done


def run_basilea_range(
    date_from: str,
    date_to: str,
    *,
    force: bool = False,
    dry_run: bool = False,
) -> list[str]:
    listing = scrape_basilea_listing()
    by_period: dict[str, int] = {}
    for aid, per in listing.items():
        if per < date_from or per > date_to:
            continue
        prev = by_period.get(per)
        if prev is None or aid > prev:
            by_period[per] = aid
    todo = sorted(by_period.items(), key=lambda t: t[0])
    log.info("Basilea range %s–%s · %d periods", date_from, date_to, len(todo))
    done = []
    for per, aid in todo:
        data = download_article_xlsx(aid)
        if not data:
            log.warning("Download failed aid=%s period=%s", aid, per)
            continue
        try:
            got = load_xlsx_bytes(
                data, force=force, dry_run=dry_run, expected_period=per
            )
        except Exception as e:
            log.exception("Parse/load failed aid=%s: %s", aid, e)
            continue
        if got:
            done.append(got)
    return done


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Chile CMF Basilea III xlsx loader")
    ap.add_argument("--article-id", type=int, help="Single CMF article ID")
    ap.add_argument("--xlsx-path", help="Local .xlsx path")
    ap.add_argument("--xlsx-url", help="Direct xlsx URL")
    ap.add_argument("--from", dest="date_from", help="YYYYMM start")
    ap.add_argument("--to", dest="date_to", help="YYYYMM end")
    ap.add_argument("--all", action="store_true", help="Load every period on the listing")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    if args.xlsx_path:
        data = Path(args.xlsx_path).read_bytes()
        load_xlsx_bytes(data, force=args.force, dry_run=args.dry_run)
        return 0
    if args.xlsx_url:
        data = _http_get(args.xlsx_url, timeout=120)
        load_xlsx_bytes(data, force=args.force, dry_run=args.dry_run)
        return 0
    if args.article_id:
        data = download_article_xlsx(args.article_id)
        if not data:
            log.error("Could not download article %s", args.article_id)
            return 1
        load_xlsx_bytes(data, force=args.force, dry_run=args.dry_run)
        return 0
    if args.date_from and args.date_to:
        run_basilea_range(
            args.date_from, args.date_to, force=args.force, dry_run=args.dry_run
        )
        return 0
    if args.all:
        listing = scrape_basilea_listing()
        periods = sorted(set(listing.values()))
        if not periods:
            return 1
        run_basilea_range(periods[0], periods[-1], force=args.force, dry_run=args.dry_run)
        return 0

    run_basilea_incremental(force=args.force, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
