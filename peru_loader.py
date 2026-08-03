#!/usr/bin/env python3
"""
peru_loader.py — ETL SBS Boletín Estadístico B-2201 → CockroachDB
LatamBanks — country='PE'

Fuente (Balance + PyG, un Excel/mes):
  https://intranet2.sbs.gob.pe/estadistica/financiera/{YYYY}/{Mes}/B-2201-{mm}{YYYY}.XLS

  Sheet 1 → tipo='b1'  (Balance General por Empresa Bancaria)
  Sheet 2 → tipo='r1'  (Estado de Ganancias y Pérdidas)

Unidad: miles de soles → monto_total en soles enteros (×1000).
ins_cod: tabla fija nombre→código (+ hash estable para bancos nuevos).

Modos:
  (sin flags)     Incremental: meses recientes aún no en carga_log
  --month AAAAMM  Carga/recarga un mes
  --all           Recarga todos los del rango
  --from / --to   Filtra AAAAMM
  --dry-run       Lista / parsea sin tocar BD
  --wipe          Borra solo country='PE'
"""
from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import ssl
import time
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.request import HTTPSHandler, Request, build_opener

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Falta openpyxl. pip install openpyxl") from e

COUNTRY = "PE"
BATCH = 500
BASE = "https://intranet2.sbs.gob.pe/estadistica/financiera"
MIN_PERIOD = "201501"
SCALE = 1000  # miles de soles → soles

MONTH_META = {
    1: ("Enero", "en"),
    2: ("Febrero", "fe"),
    3: ("Marzo", "ma"),
    4: ("Abril", "ab"),
    5: ("Mayo", "my"),
    6: ("Junio", "jn"),
    7: ("Julio", "jl"),
    8: ("Agosto", "ag"),
    9: ("Setiembre", "se"),
    10: ("Octubre", "oc"),
    11: ("Noviembre", "no"),
    12: ("Diciembre", "di"),
}

# Códigos estables (sidebar / logos). No hay BTG en Banca Múltiple SBS.
BANK_CODE_BY_NORM = {
    "banco bbva peru": 1,
    "bancom": 2,
    "banco de credito del peru": 3,
    "banco pichincha": 4,
    "banco interamericano de finanzas": 5,  # BanBif
    "scotiabank peru": 6,
    "citibank": 7,
    "interbank": 8,
    "mibanco": 9,
    "banco gnb": 10,
    "banco falabella peru": 11,
    "banco santander peru": 12,
    "banco ripley": 13,
    "alfin banco": 14,
    "banco icbc": 15,
    "bank of china": 16,
    "banco bci peru": 17,
    "compartamos banco": 18,
    "santander consumer bank": 19,
}

# Agregados / duplicados a excluir del universo dashboard
EXCLUDE_NAME_RE = re.compile(
    r"(total\s+banca|sucursales\s+en\s+el\s+exterior|incluye\s+sucursales)",
    re.I,
)

# Labels SBS → cuenta canónica (KPIs). Match por slug normalizado / contains.
CANONICAL_B1 = {
    "TOTAL_ACTIVO": "TOTAL_ACTIVO",
    "CREDITOS_NETOS_DE_PROVISIONES_Y_DE_INGRESOS_NO_DEVENGADOS": "CREDITOS_NETOS",
    "OBLIGACIONES_CON_EL_PUBLICO": "OBLIGACIONES_PUBLICO",
    "TOTAL_PASIVO": "TOTAL_PASIVO",
    "PATRIMONIO": "PATRIMONIO",
    "DEPOSITOS_A_LA_VISTA": "DEPOSITOS_VISTA",
    "DEPOSITOS_DE_AHORRO": "DEPOSITOS_AHORRO",
    "DEPOSITOS_A_PLAZO": "DEPOSITOS_PLAZO",
    "DISPONIBLE": "DISPONIBLE",
    "INVERSIONES_NETAS_DE_PROVISIONES": "INVERSIONES_NETAS",
    "RESULTADO_NETO_DEL_EJERCICIO": "RESULTADO_NETO_PATRIMONIO",
}

CANONICAL_R1 = {
    "INGRESOS_FINANCIEROS": "INGRESOS_FINANCIEROS",
    "GASTOS_FINANCIEROS": "GASTOS_FINANCIEROS",
    "MARGEN_FINANCIERO_BRUTO": "MARGEN_FINANCIERO_BRUTO",
    "PROVISIONES_PARA_CREDITOS_DIRECTOS": "PROVISIONES_CREDITOS",
    "MARGEN_FINANCIERO_NETO": "MARGEN_FINANCIERO_NETO",
    "INGRESOS_POR_SERVICIOS_FINANCIEROS": "INGRESOS_SERVICIOS",
    "GASTOS_POR_SERVICIOS_FINANCIEROS": "GASTOS_SERVICIOS",
    "MARGEN_OPERACIONAL": "MARGEN_OPERACIONAL",
    "GASTOS_ADMINISTRATIVOS": "GASTOS_ADMINISTRATIVOS",
    "MARGEN_OPERACIONAL_NETO": "MARGEN_OPERACIONAL_NETO",
    "RESULTADO_ANTES_DE_IMPUESTO_A_LA_RENTA": "RESULTADO_ANTES_IMPUESTO",
    "IMPUESTO_A_LA_RENTA": "IMPUESTO_RENTA",
    "RESULTADO_NETO_DEL_EJERCICIO": "RESULTADO_NETO",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("peru_loader")

_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE
_OPENER = build_opener(HTTPSHandler(context=_SSL_CTX))


def fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def slug_label(label: str) -> str:
    s = fold(label).upper().replace(" ", "_")
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    # quitar notas al pie tipo "1/" al final
    s = re.sub(r"_?\d+$", "", s) if re.search(r"_\d+$", s) and "SUBORDINADAS" in s else s
    s = re.sub(r"_1$", "", s)
    return s


def period_to_ym(periodo: str) -> tuple[int, int]:
    return int(periodo[:4]), int(periodo[4:6])


def ym_to_period(y: int, m: int) -> str:
    return f"{y}{m:02d}"


def iter_periods(start: str, end: str) -> list[str]:
    y, m = period_to_ym(start)
    ye, me = period_to_ym(end)
    out = []
    while (y, m) <= (ye, me):
        out.append(ym_to_period(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def file_url(periodo: str) -> str:
    y, m = period_to_ym(periodo)
    folder, suf = MONTH_META[m]
    return f"{BASE}/{y}/{folder}/B-2201-{suf}{y}.XLS"


def http_bytes(url: str, retries: int = 3, backoff: float = 4.0) -> bytes:
    last = None
    for attempt in range(1, retries + 1):
        try:
            req = Request(url, headers={"User-Agent": "LatamBanksPE/1.0"})
            with _OPENER.open(req, timeout=120) as r:
                return r.read()
        except (HTTPError, URLError, TimeoutError, ssl.SSLError) as e:
            last = e
            log.warning("intento %d/%d fallo %s: %s", attempt, retries, url, e)
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise last


def index_exists(periodo: str) -> bool:
    try:
        http_bytes(file_url(periodo), retries=1)
        return True
    except Exception:
        return False


def discover_available_periods(start: str = MIN_PERIOD, end: str | None = None) -> list[str]:
    if end is None:
        from datetime import date

        today = date.today()
        end = ym_to_period(today.year, today.month)
    found = []
    for p in iter_periods(max(start, MIN_PERIOD), end):
        if index_exists(p):
            found.append(p)
            log.info("catálogo: %s OK", p)
    log.info(
        "Catálogo PE: %d meses (%s .. %s)",
        len(found),
        found[0] if found else "—",
        found[-1] if found else "—",
    )
    return found


def recent_candidate_periods(n_months: int = 4) -> list[str]:
    from datetime import date

    today = date.today()
    y, m = today.year, today.month
    out = []
    for _ in range(n_months):
        out.append(ym_to_period(y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))


def bank_code(name: str) -> int:
    n = fold(name).replace("\n", " ")
    n = re.sub(r"\s+", " ", n).strip()
    if n in BANK_CODE_BY_NORM:
        return BANK_CODE_BY_NORM[n]
    # aliases cortos
    aliases = {
        "bbva": 1,
        "bcp": 3,
        "banbif": 5,
        "scotiabank": 6,
        "falabella": 11,
        "santander": 12,
        "ripley": 13,
        "alfin": 14,
        "icbc": 15,
        "bci": 17,
        "compartamos": 18,
    }
    for k, v in aliases.items():
        if k in n:
            return v
    h = int(hashlib.md5(n.encode()).hexdigest()[:6], 16)
    return 1000 + (h % 8000)


def is_excluded_bank(name: str) -> bool:
    return bool(EXCLUDE_NAME_RE.search(fold(name)))


def detect_banks(sh) -> list[tuple[int, str, int]]:
    """[(ins_cod, nombre, total_col), ...] excluyendo agregados."""
    out = []
    seen = set()
    for c in range(1, (sh.max_column or 0) + 1):
        name = sh.cell(6, c).value
        if not name:
            continue
        ns = re.sub(r"\s+", " ", str(name).replace("\n", " ")).strip()
        if not ns or fold(ns) in ("activo", "pasivo"):
            continue
        if is_excluded_bank(ns):
            continue
        if sh.cell(7, c).value != "MN":
            continue
        tot_col = c + 2
        code = bank_code(ns)
        if code in seen:
            # colisión rara: desplazar
            code = code + 100
        seen.add(code)
        out.append((code, ns, tot_col))
    return out


def canonical_cuenta(label: str, tipo: str) -> tuple[str, str] | None:
    lab = str(label or "").strip()
    if not lab:
        return None
    # saltar encabezados / notas
    low = fold(lab)
    if low.startswith("(en miles") or low.startswith("tipo de cambio") or low.startswith("balance general"):
        return None
    if low.startswith("estado de ganancias"):
        return None
    if low in ("activo", "pasivo"):
        return None
    if re.match(r"^\d+/", low):
        return None

    slug = slug_label(lab)
    if not slug or len(slug) < 2:
        return None

    table = CANONICAL_B1 if tipo == "b1" else CANONICAL_R1
    if slug in table:
        return table[slug], lab

    # contains heuristics for KPI rows with slight wording drift
    if tipo == "b1":
        if slug == "TOTAL_ACTIVO" or slug.startswith("TOTAL_ACTIVO"):
            return "TOTAL_ACTIVO", lab
        if "CREDITOS_NETOS" in slug and "PROVISIONES" in slug:
            return "CREDITOS_NETOS", lab
        if slug == "OBLIGACIONES_CON_EL_PUBLICO":
            return "OBLIGACIONES_PUBLICO", lab
        if slug == "PATRIMONIO":
            return "PATRIMONIO", lab
        if slug == "TOTAL_PASIVO":
            return "TOTAL_PASIVO", lab
    if tipo == "r1":
        if "RESULTADO_NETO_DEL_EJERCICIO" in slug or slug == "RESULTADO_NETO_DEL_EJERCICIO":
            return "RESULTADO_NETO", lab

    return slug, lab


def _cell_num(sh, r: int, c: int) -> int:
    v = sh.cell(r, c).value
    if v in ("", None):
        return 0
    try:
        return int(round(float(v) * SCALE))
    except (TypeError, ValueError):
        return 0


def parse_sheet(sh, tipo: str, periodo: str, banks: list[tuple[int, str, int]]):
    rows = []
    plan: dict[str, str] = {}
    seen_cuentas: set[str] = set()
    # Prefer label from col 1; fallback col 13 (panel derecho del layout SBS).
    # Primera aparición gana (ej. Depósitos a la Vista bajo Obligaciones con el
    # Público, no el bloque del sistema financiero).
    for r in range(1, (sh.max_row or 0) + 1):
        label = sh.cell(r, 1).value or sh.cell(r, 13).value
        parsed = canonical_cuenta(label, tipo)
        if not parsed:
            continue
        cuenta, desc = parsed
        if cuenta in seen_cuentas:
            # KPIs canónicos: primera aparición (p.ej. depósitos del público).
            # Detalle: desambiguar si el label se repite (activo vs pasivo).
            short = set(CANONICAL_B1.values()) | set(CANONICAL_R1.values())
            if cuenta in short:
                continue
            n = 2
            while f"{cuenta}_{n}" in seen_cuentas:
                n += 1
            cuenta = f"{cuenta}_{n}"
        seen_cuentas.add(cuenta)
        plan[cuenta] = desc
        for ins_cod, _name, tot_col in banks:
            val = _cell_num(sh, r, tot_col)
            rows.append((COUNTRY, periodo, tipo, ins_cod, cuenta, 0, 0, 0, 0, val))
    return rows, plan


def parse_workbook(data: bytes, periodo: str):
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    if "1" not in wb.sheetnames:
        raise RuntimeError(f"Sheet '1' ausente en {periodo}: {wb.sheetnames}")
    sh1 = wb["1"]
    banks = detect_banks(sh1)
    if not banks:
        raise RuntimeError(f"Sin bancos detectados en {periodo}")
    rows_b1, plan = parse_sheet(sh1, "b1", periodo, banks)
    rows_r1 = []
    if "2" in wb.sheetnames:
        # re-detect banks on sheet 2 (misma geometría habitual)
        banks2 = detect_banks(wb["2"]) or banks
        # alinear por nombre si difiere el orden
        by_name = {fold(n): (c, n, t) for c, n, t in banks2}
        aligned = []
        for code, name, tot in banks:
            hit = by_name.get(fold(name))
            if hit:
                aligned.append((code, name, hit[2]))
            else:
                aligned.append((code, name, tot))
        rows_r1, plan2 = parse_sheet(wb["2"], "r1", periodo, aligned)
        plan.update(plan2)
    inst = [(COUNTRY, code, name) for code, name, _ in banks]
    return inst, rows_b1 + rows_r1, plan


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def get_db_url() -> str:
    url = os.environ.get("COCKROACH_URL")
    if not url:
        raise SystemExit("Falta COCKROACH_URL (.env o entorno).")
    return url


def connect():
    import psycopg2

    return psycopg2.connect(get_db_url())


def upsert(conn, table, cols, updates, conflict, rows):
    import psycopg2.extras

    if not rows:
        return
    cur = conn.cursor()
    sql = (
        f"INSERT INTO {table} ({','.join(cols)}) VALUES %s "
        f"ON CONFLICT ({','.join(conflict)}) DO UPDATE SET "
        + ", ".join(f"{c}=EXCLUDED.{c}" for c in updates)
    )
    for i in range(0, len(rows), BATCH):
        psycopg2.extras.execute_values(cur, sql, rows[i : i + BATCH])
    conn.commit()


def wipe_pe(conn):
    cur = conn.cursor()
    for t in ("datos_financieros", "instituciones", "plan_cuentas", "carga_log"):
        cur.execute(f"DELETE FROM {t} WHERE country=%s", (COUNTRY,))
    conn.commit()
    log.info("Perú borrado (solo country='PE').")


def get_loaded_periods(conn) -> set[str]:
    cur = conn.cursor()
    cur.execute(
        "SELECT periodo FROM carga_log WHERE country=%s AND estado IN ('ok','alerta_esquema')",
        (COUNTRY,),
    )
    return {r[0] for r in cur.fetchall()}


def load_month(conn, periodo: str) -> None:
    from schema_guard import detect_schema_changes, get_known_accounts, record_schema_result

    url = file_url(periodo)
    log.info("dt=%s descargando %s", periodo, url)
    data = http_bytes(url)
    inst_rows, all_rows, plan = parse_workbook(data, periodo)
    if not all_rows:
        raise RuntimeError(f"Sin filas de datos para {periodo}")

    known = get_known_accounts(conn, COUNTRY)
    incoming = {r[4] for r in all_rows}
    report = detect_schema_changes(COUNTRY, periodo, incoming, known)

    upsert(
        conn,
        "instituciones",
        ["country", "codigo", "razon_social"],
        ["razon_social"],
        ["country", "codigo"],
        inst_rows,
    )
    upsert(
        conn,
        "plan_cuentas",
        ["country", "cuenta", "descripcion"],
        ["descripcion"],
        ["country", "cuenta"],
        [(COUNTRY, c, d) for c, d in sorted(plan.items())],
    )
    upsert(
        conn,
        "datos_financieros",
        [
            "country",
            "periodo",
            "tipo",
            "ins_cod",
            "cuenta",
            "monto_clp",
            "monto_uf",
            "monto_tc",
            "monto_ext",
            "monto_total",
        ],
        ["monto_total"],
        ["country", "periodo", "tipo", "ins_cod", "cuenta"],
        all_rows,
    )
    upsert(
        conn,
        "carga_log",
        ["country", "periodo", "archivos_procesados", "estado"],
        ["archivos_procesados", "estado"],
        ["country", "periodo"],
        [(COUNTRY, periodo, len(inst_rows), "ok")],
    )
    record_schema_result(conn, COUNTRY, periodo, report)
    log.info(
        "dt=%s OK: %d bancos, %d filas, %d cuentas (schema=%s)",
        periodo,
        len(inst_rows),
        len(all_rows),
        len(plan),
        report.get("status"),
    )


def main(argv: Iterable[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="ETL SBS B-2201 Perú → CockroachDB")
    ap.add_argument("--wipe", action="store_true")
    ap.add_argument("--month", help="AAAAMM puntual")
    ap.add_argument("--all", action="store_true", help="Recargar todo el rango")
    ap.add_argument("--from", dest="from_dt", default=MIN_PERIOD)
    ap.add_argument("--to", dest="to_dt", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--skip-discover",
        action="store_true",
        help="No sondear URLs: usa rango calendario (fallará en meses sin archivo)",
    )
    args = ap.parse_args(list(argv) if argv is not None else None)

    if args.month:
        periods = [args.month]
    elif args.all or args.to_dt or (args.from_dt and args.from_dt != MIN_PERIOD):
        periods = (
            iter_periods(args.from_dt, args.to_dt)
            if args.skip_discover and args.to_dt
            else discover_available_periods(args.from_dt, args.to_dt)
        )
    else:
        periods = [p for p in recent_candidate_periods(10) if index_exists(p)]

    if args.from_dt:
        periods = [p for p in periods if p >= args.from_dt]
    if args.to_dt:
        periods = [p for p in periods if p <= args.to_dt]

    if args.dry_run:
        print(f"Meses objetivo: {len(periods)}")
        for p in periods:
            print(f"  {p}  {file_url(p)}")
        if periods:
            data = http_bytes(file_url(periods[-1]))
            inst, rows, plan = parse_workbook(data, periods[-1])
            print(f"Bancos ({periods[-1]}): {sorted((r[1], r[2]) for r in inst)}")
            by = {}
            for r in rows:
                if r[3] == 3 and r[4] in (
                    "TOTAL_ACTIVO",
                    "CREDITOS_NETOS",
                    "OBLIGACIONES_PUBLICO",
                    "PATRIMONIO",
                    "RESULTADO_NETO",
                ):
                    by[(r[2], r[4])] = r[9]
            print("BCP sample KPIs (soles):", by)
            print(f"filas={len(rows)} plan={len(plan)}")
        return 0

    conn = connect()
    try:
        if args.wipe:
            wipe_pe(conn)

        if not args.month and not args.all:
            loaded = get_loaded_periods(conn)
            periods = [p for p in periods if p not in loaded]
            if not periods:
                log.info("Perú al día. Nada nuevo.")
                return 0

        log.info(
            "Cargando %d mes(es): %s%s",
            len(periods),
            periods[:6],
            "…" if len(periods) > 6 else "",
        )
        for p in periods:
            load_month(conn, p)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
