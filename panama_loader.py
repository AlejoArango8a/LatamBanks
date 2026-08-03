#!/usr/bin/env python3
"""
panama_loader.py — ETL Superintendencia de Bancos de Panamá (SBP) → CockroachDB
LatamBanks — country='PA'

Fuente (Excel individual por banco, miles de balboas = USD×1000):
  .../reportes_estadisticos/{YYYY}/{MM}/balance_individual_por_banco/RE-BALANCE-BANCO-en-{Slug}.xlsx
  .../reportes_estadisticos/{YYYY}/{MM}/estado_de_resultado_individual_por_banco/RE-ESTADO-BANCO-en-{Slug}.xlsx

Universo: Bancos Oficiales + Licencia General + Licencia Internacional
  (usar --domestic-only para excluir Licencia Internacional).
PyG: RESULTADO_NETO se guarda como YTD (suma de flujos mensuales ene→mes).

Modos: --month / --all / --from / --to / --dry-run / --wipe / --domestic-only
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import ssl
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.request import HTTPSHandler, Request, build_opener

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Falta openpyxl. pip install openpyxl") from e

COUNTRY = "PA"
BATCH = 500
MIN_PERIOD = "202001"
SCALE = 1000  # miles de balboas → USD enteros
BASE = (
    "https://www.superbancos.gob.pa/documentos/financiera_y_estadistica/"
    "reportes_estadisticos"
)
# Índices anuales Drupal (scrape de slugs). Ampliar cuando SBP publique nodos nuevos.
YEAR_INDEX_NODES = {
    2024: 1228,
    2025: 1429,
    2026: 1669,
}

MONTH_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
# SBP a veces usa Setiembre
MONTH_ALIASES = {"Setiembre": 9, "Septiembre": 9}

B1_LABELS = {
    "TOTAL DE ACTIVOS": "TOTAL_ACTIVO",
    "CARTERA CREDITICIA": "CARTERA_CREDITICIA",
    "DEPOSITOS": "DEPOSITOS",
    "OBLIGACIONES": "OBLIGACIONES",
    "OTROS PASIVOS": "OTROS_PASIVOS",
    "PATRIMONIO": "PATRIMONIO",
}
R1_LABELS = {
    "Utilidad del Periodo": "RESULTADO_NETO",
    "Utilidad antes de Provisiones": "UTILIDAD_ANTES_PROVISIONES",
}

# Catálogo fallback Oficiales + Licencia General (dic-2025). Si un slug 404 → skip.
FALLBACK_SLUGS_DOMESTIC = [
    "Nacional", "Cajahorros",
    "Allbank", "Atlas", "Bac", "balboa", "Aliado", "Azteca", "BBVA", "Davivienda",
    "Bogotasuc", "Delta", "FICOGRAL", "GTContinental", "General", "Bicsa",
    "Hipotecaria", "Lafise", "Bladex", "BancoPanama", "Banvivienda", "Pichincha",
    "BcoPrival", "Bancolopanama", "BBPbank", "Banescosa", "Banisi", "Banistmosa",
    "China", "BCTBankIntSA", "bibank", "Canal", "Citi", "Credicorp", "fpbbank",
    "Global", "ICBCL", "Korea", "ICBC", "MercantilPanama", "Metrobank", "MMG",
    "Multibanksub", "Pacificogral", "Stgeorge", "Scotia", "Tower", "Unibank",
]

# Licencia Internacional (índice SBP 2026). Incluye Banco de Occidente (Panamá).
FALLBACK_SLUGS_INTERNATIONAL = [
    "ANDBANC", "Atlantic", "Austrobank", "ASBbank", "BBVA", "BCreditoAndorra",
    "Davint", "Bogotaint", "CreditoPeru", "Argentina", "Occidente", "Bancolombia",
    "BHDint", "BPR", "GNB", "Inteligo", "Unionbank", "Itau", "Popular",
]

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("panama_loader")

_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE
_OPENER = build_opener(HTTPSHandler(context=_SSL_CTX))
_UA = "Mozilla/5.0 (compatible; LatamBanksPA/1.0)"


def http_bytes(url: str, retries: int = 3, backoff: float = 2.5) -> bytes:
    last = None
    for attempt in range(1, retries + 1):
        try:
            req = Request(url, headers={"User-Agent": _UA})
            with _OPENER.open(req, timeout=90) as r:
                return r.read()
        except (HTTPError, URLError, TimeoutError, ssl.SSLError) as e:
            last = e
            if isinstance(e, HTTPError) and e.code == 404:
                raise
            log.warning("intento %d/%d %s: %s", attempt, retries, url, e)
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise last


def http_exists(url: str) -> bool:
    try:
        req = Request(url, headers={"User-Agent": _UA, "Range": "bytes=0-64"})
        with _OPENER.open(req, timeout=30) as r:
            r.read(8)
            return True
    except Exception:
        return False


def period_to_ym(periodo: str) -> tuple[int, int]:
    return int(periodo[:4]), int(periodo[4:6])


def ym_to_period(y: int, m: int) -> str:
    return f"{y}{m:02d}"


def iter_periods(start: str, end: str) -> list[str]:
    y, m = period_to_ym(start)
    ey, em = period_to_ym(end)
    out = []
    while (y, m) <= (ey, em):
        out.append(ym_to_period(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def balance_url(periodo: str, slug: str) -> str:
    y, m = period_to_ym(periodo)
    return (
        f"{BASE}/{y}/{m:02d}/balance_individual_por_banco/"
        f"RE-BALANCE-BANCO-en-{slug}.xlsx"
    )


def estado_url(periodo: str, slug: str) -> str:
    y, m = period_to_ym(periodo)
    return (
        f"{BASE}/{y}/{m:02d}/estado_de_resultado_individual_por_banco/"
        f"RE-ESTADO-BANCO-en-{slug}.xlsx"
    )


def discover_periods(from_dt: str, to_dt: str | None) -> list[str]:
    end = to_dt or ym_to_period(*time.gmtime()[:2])
    cands = [p for p in iter_periods(from_dt, end) if p >= MIN_PERIOD]
    # Probe with Nacional (always present for domestic months)
    ok = []
    for p in cands:
        if http_exists(balance_url(p, "Nacional")):
            ok.append(p)
        else:
            # allow gaps without aborting
            pass
    return ok


def scrape_slugs_for_year(year: int, include_international: bool = True) -> list[str] | None:
    node = YEAR_INDEX_NODES.get(year)
    if not node:
        return None
    url = f"https://www.superbancos.gob.pa/node/{node}"
    try:
        html = http_bytes(url).decode("utf-8", errors="ignore")
    except Exception as e:
        log.warning("no se pudo scrape index %s: %s", url, e)
        return None

    sec_pat = list(
        re.finditer(
            r"<b>\s*(Bancos Oficiales|Bancos de Licencia General|Bancos de Licencia Internacional)\s*</b>",
            html,
            re.I,
        )
    )

    def section_at(pos: int) -> str:
        cur = ""
        for m in sec_pat:
            if m.start() <= pos:
                cur = m.group(1).strip().lower()
            else:
                break
        return cur

    want = []
    for m in re.finditer(
        rf"/{year}/(\d{{2}})/balance_individual_por_banco/RE-BALANCE-BANCO-en-([A-Za-z0-9_-]+)\.xlsx",
        html,
    ):
        sec = section_at(m.start())
        if "internacional" in sec and not include_international:
            continue
        if "oficial" in sec or "general" in sec or (include_international and "internacional" in sec):
            want.append(m.group(2))
    # unique preserve order
    seen = set()
    out = []
    for s in want:
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out or None


def slugs_for_period(periodo: str, include_international: bool = True) -> list[str]:
    y, _ = period_to_ym(periodo)
    scraped = scrape_slugs_for_year(y, include_international=include_international)
    if scraped:
        return scraped
    out = list(FALLBACK_SLUGS_DOMESTIC)
    if include_international:
        seen = set(out)
        for s in FALLBACK_SLUGS_INTERNATIONAL:
            if s not in seen:
                out.append(s)
                seen.add(s)
    return out


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _month_num(label) -> int | None:
    if label is None:
        return None
    if hasattr(label, "month"):
        return int(label.month)
    t = _norm(label)
    if t in MONTH_ALIASES:
        return MONTH_ALIASES[t]
    for n, name in MONTH_ES.items():
        if name.lower() == t.lower():
            return n
    return None


def _sheet_grid(data: bytes):
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _find_month_col(header_row: list, target_mm: int, prefer_last: bool = True) -> int | None:
    hits = []
    for idx, cell in enumerate(header_row):
        if _month_num(cell) == target_mm:
            hits.append(idx)
    if not hits:
        return None
    return hits[-1] if prefer_last else hits[0]


def parse_balance(data: bytes, periodo: str) -> tuple[int, str, dict[str, int], dict[str, str]]:
    """Returns (ins_cod, name, {cuenta: monto}, plan)."""
    y, m = period_to_ym(periodo)
    rows = _sheet_grid(data)
    if len(rows) < 12:
        raise RuntimeError("balance demasiado corto")
    name = _norm(rows[2][0] if len(rows) > 2 else "")
    code_raw = _norm(rows[7][0] if len(rows) > 7 else "")
    if not code_raw.isdigit():
        raise RuntimeError(f"código SBP inválido en A8: {code_raw!r}")
    ins_cod = int(code_raw)
    header = rows[9] if len(rows) > 9 else []
    col = _find_month_col(header, m, prefer_last=True)
    if col is None:
        raise RuntimeError(f"sin columna mes {m} en balance {name}")

    vals: dict[str, int] = {}
    plan: dict[str, str] = {}
    for r in rows[10:]:
        label = _norm(r[0] if r else "")
        if label not in B1_LABELS:
            continue
        cuenta = B1_LABELS[label]
        raw = r[col] if col < len(r) else None
        try:
            num = float(raw or 0)
        except (TypeError, ValueError):
            num = 0.0
        vals[cuenta] = int(round(num * SCALE))
        plan[cuenta] = label

    # Componer pasivo total
    if all(k in vals for k in ("DEPOSITOS", "OBLIGACIONES", "OTROS_PASIVOS")):
        vals["TOTAL_PASIVO"] = vals["DEPOSITOS"] + vals["OBLIGACIONES"] + vals["OTROS_PASIVOS"]
        plan["TOTAL_PASIVO"] = "TOTAL PASIVO (depósitos+obligaciones+otros)"
    return ins_cod, name, vals, plan


def parse_estado_ytd(data: bytes, periodo: str) -> tuple[dict[str, int], dict[str, str]]:
    """PyG: RESULTADO_NETO as YTD sum of monthly flows ene→target month."""
    y, m = period_to_ym(periodo)
    rows = _sheet_grid(data)
    header = rows[9] if len(rows) > 9 else []
    # Map month -> column index for CURRENT year months only.
    # Row 9 often has year on first col of group; row 10 has month names.
    # Prior-year December is the first Diciembre; current-year months follow.
    year_row = rows[8] if len(rows) > 8 else []
    # Find columns belonging to current year: after the first December (prior),
    # or columns under year==y in year_row.
    month_cols: dict[int, int] = {}
    seen_prior_dec = False
    for idx, cell in enumerate(header):
        mm = _month_num(cell)
        if mm is None:
            continue
        yr = year_row[idx] if idx < len(year_row) else None
        # Heuristic: first December is prior year annual; subsequent months current
        if mm == 12 and not seen_prior_dec:
            # check if this is prior: if year cell is y-1 or first dec
            if yr == y - 1 or yr is None or yr == y - 1.0:
                seen_prior_dec = True
                continue
            # if year cell says current year, it's current Dec
            if yr == y or yr == float(y):
                month_cols[12] = idx
                continue
            seen_prior_dec = True
            continue
        if mm == 12 and seen_prior_dec:
            month_cols[12] = idx
            continue
        # Non-December: assign to current year after we've passed prior Dec,
        # or if year_row marks current year.
        if seen_prior_dec or yr == y or yr == float(y):
            month_cols[mm] = idx
            if mm != 12:
                seen_prior_dec = True

    # Fallback: if heuristic failed, take last occurrence of each month 1..m
    if not month_cols:
        for idx, cell in enumerate(header):
            mm = _month_num(cell)
            if mm and mm <= m:
                month_cols[mm] = idx

    cols_ytd = [month_cols[mm] for mm in range(1, m + 1) if mm in month_cols]
    # Also try ACUMULADO column for December
    acum_col = None
    for idx, cell in enumerate(header):
        if _norm(cell).upper() == "ACUMULADO":
            acum_col = idx

    vals: dict[str, int] = {}
    plan: dict[str, str] = {}
    for r in rows[10:]:
        label = _norm(r[0] if r else "")
        if label not in R1_LABELS:
            continue
        cuenta = R1_LABELS[label]
        if cuenta == "RESULTADO_NETO" and m == 12 and acum_col is not None:
            try:
                num = float(r[acum_col] or 0)
            except (TypeError, ValueError):
                num = 0.0
        else:
            num = 0.0
            for c in cols_ytd:
                try:
                    num += float(r[c] or 0)
                except (TypeError, ValueError):
                    pass
        vals[cuenta] = int(round(num * SCALE))
        plan[cuenta] = f"{label} (YTD)"
    return vals, plan


def fetch_bank(periodo: str, slug: str) -> tuple[str, tuple | None, str | None]:
    """Download+parse one bank. Returns (slug, payload|None, error|None).
    payload = (ins_cod, name, b1_vals, r1_vals, plan)
    """
    try:
        bal = http_bytes(balance_url(periodo, slug))
        ins_cod, name, b1, plan_b = parse_balance(bal, periodo)
        r1: dict[str, int] = {}
        plan_r: dict[str, str] = {}
        try:
            est = http_bytes(estado_url(periodo, slug))
            r1, plan_r = parse_estado_ytd(est, periodo)
        except HTTPError as e:
            if e.code != 404:
                raise
            log.warning("dt=%s slug=%s sin estado de resultados", periodo, slug)
        plan = {**plan_b, **plan_r}
        return slug, (ins_cod, name, b1, r1, plan), None
    except HTTPError as e:
        if e.code == 404:
            return slug, None, "404"
        return slug, None, str(e)
    except Exception as e:
        return slug, None, str(e)


def build_month_rows(periodo: str, include_international: bool = True, workers: int = 8):
    slugs = slugs_for_period(periodo, include_international=include_international)
    inst = []
    data_rows = []
    plan: dict[str, str] = {}
    ok = 0
    skip = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(fetch_bank, periodo, s) for s in slugs]
        for fut in as_completed(futs):
            slug, payload, err = fut.result()
            if err == "404" or payload is None:
                skip += 1
                continue
            if err:
                log.warning("dt=%s slug=%s error: %s", periodo, slug, err)
                skip += 1
                continue
            ins_cod, name, b1, r1, p = payload
            inst.append((COUNTRY, ins_cod, name))
            plan.update(p)
            for cuenta, monto in b1.items():
                data_rows.append(
                    (COUNTRY, periodo, "b1", ins_cod, cuenta, 0, 0, 0, 0, monto)
                )
            for cuenta, monto in r1.items():
                data_rows.append(
                    (COUNTRY, periodo, "r1", ins_cod, cuenta, 0, 0, 0, 0, monto)
                )
            ok += 1
    # dedupe instituciones by code (last name wins)
    by_code = {c: (COUNTRY, c, n) for _, c, n in inst}
    return list(by_code.values()), data_rows, plan, ok, skip


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def get_db_url() -> str:
    url = os.environ.get("COCKROACH_URL")
    if not url:
        raise SystemExit("Falta COCKROACH_URL (.env o entorno).")
    return url


def connect():
    import psycopg2

    return psycopg2.connect(get_db_url())


def upsert(conn, table, cols, updates, conflict, rows):
    import psycopg2.extras

    if not rows:
        return
    cur = conn.cursor()
    sql = (
        f"INSERT INTO {table} ({','.join(cols)}) VALUES %s "
        f"ON CONFLICT ({','.join(conflict)}) DO UPDATE SET "
        + ", ".join(f"{c}=EXCLUDED.{c}" for c in updates)
    )
    for i in range(0, len(rows), BATCH):
        psycopg2.extras.execute_values(cur, sql, rows[i : i + BATCH])
    conn.commit()


def wipe_pa(conn):
    cur = conn.cursor()
    for t in ("datos_financieros", "instituciones", "plan_cuentas", "carga_log"):
        cur.execute(f"DELETE FROM {t} WHERE country=%s", (COUNTRY,))
    conn.commit()
    log.info("Panamá borrado (solo country='PA').")


def get_loaded_periods(conn) -> set[str]:
    cur = conn.cursor()
    cur.execute(
        "SELECT periodo FROM carga_log WHERE country=%s AND estado IN ('ok','alerta_esquema')",
        (COUNTRY,),
    )
    return {r[0] for r in cur.fetchall()}


def load_month(conn, periodo: str, include_international: bool = True) -> None:
    from schema_guard import detect_schema_changes, get_known_accounts, record_schema_result

    scope = "Oficiales+General+Internacional" if include_international else "Oficiales+General"
    log.info("dt=%s cargando SBP %s…", periodo, scope)
    inst_rows, all_rows, plan, ok, skip = build_month_rows(
        periodo, include_international=include_international
    )
    if not all_rows:
        raise RuntimeError(f"Sin filas de datos para {periodo} (ok={ok} skip={skip})")

    known = get_known_accounts(conn, COUNTRY)
    incoming = {r[4] for r in all_rows}
    report = detect_schema_changes(COUNTRY, periodo, incoming, known)

    upsert(
        conn,
        "instituciones",
        ["country", "codigo", "razon_social"],
        ["razon_social"],
        ["country", "codigo"],
        inst_rows,
    )
    upsert(
        conn,
        "plan_cuentas",
        ["country", "cuenta", "descripcion"],
        ["descripcion"],
        ["country", "cuenta"],
        [(COUNTRY, c, d) for c, d in sorted(plan.items())],
    )
    upsert(
        conn,
        "datos_financieros",
        [
            "country",
            "periodo",
            "tipo",
            "ins_cod",
            "cuenta",
            "monto_clp",
            "monto_uf",
            "monto_tc",
            "monto_ext",
            "monto_total",
        ],
        ["monto_total"],
        ["country", "periodo", "tipo", "ins_cod", "cuenta"],
        all_rows,
    )
    upsert(
        conn,
        "carga_log",
        ["country", "periodo", "archivos_procesados", "estado"],
        ["archivos_procesados", "estado"],
        ["country", "periodo"],
        [(COUNTRY, periodo, ok, "ok")],
    )
    record_schema_result(conn, COUNTRY, periodo, report)
    log.info(
        "dt=%s OK: %d bancos, %d filas, skip=%d (schema=%s)",
        periodo,
        ok,
        len(all_rows),
        skip,
        report.get("status"),
    )


def main(argv: Iterable[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="ETL SBP Panamá → CockroachDB")
    ap.add_argument("--wipe", action="store_true")
    ap.add_argument("--month", help="AAAAMM puntual")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--from", dest="from_dt", default=MIN_PERIOD)
    ap.add_argument("--to", dest="to_dt", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--include-international",
        action="store_true",
        default=True,
        help="Incluir Licencia Internacional (default: sí)",
    )
    ap.add_argument(
        "--domestic-only",
        action="store_true",
        help="Solo Oficiales + Licencia General (excluye Internacional)",
    )
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args(list(argv) if argv is not None else None)
    include_international = not args.domestic_only
    # Allow explicit --include-international for backwards compatibility with older cron scripts.
    if args.include_international and args.domestic_only:
        log.warning("--domestic-only gana sobre --include-international")
        include_international = False

    if args.month:
        periods = [args.month]
    else:
        periods = discover_periods(args.from_dt, args.to_dt)
        if args.from_dt:
            periods = [p for p in periods if p >= args.from_dt]
        if args.to_dt:
            periods = [p for p in periods if p <= args.to_dt]
        if not args.all and not args.to_dt and args.from_dt == MIN_PERIOD and not args.month:
            # incremental default: last few available
            periods = periods[-10:]

    if args.dry_run:
        print(f"Meses objetivo: {len(periods)}")
        for p in periods[:12]:
            print(f"  {p}")
        if len(periods) > 12:
            print("  …")
        if periods:
            p = periods[-1]
            slugs = slugs_for_period(p, include_international)
            print(f"Slugs ({p}): {len(slugs)} → {slugs[:8]}…")
            # sample Nacional + General + Occidente (intl)
            for slug in ("Nacional", "General", "Occidente"):
                s, payload, err = fetch_bank(p, slug)
                if err or not payload:
                    print(f"  {slug}: ERROR {err}")
                    continue
                code, name, b1, r1, _ = payload
                print(f"  {slug}: {code} {name}")
                print("   b1", {k: b1.get(k) for k in ("TOTAL_ACTIVO", "PATRIMONIO", "DEPOSITOS")})
                print("   r1", r1)
        return 0

    conn = connect()
    try:
        if args.wipe:
            wipe_pa(conn)
        if not args.month and not args.all:
            loaded = get_loaded_periods(conn)
            periods = [p for p in periods if p not in loaded]
            if not periods:
                log.info("Panamá al día. Nada nuevo.")
                return 0
        log.info(
            "Cargando %d mes(es) (intl=%s): %s%s",
            len(periods),
            include_international,
            periods[:6],
            "…" if len(periods) > 6 else "",
        )
        for p in periods:
            load_month(conn, p, include_international=include_international)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
