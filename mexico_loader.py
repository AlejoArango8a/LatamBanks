#!/usr/bin/env python3
"""
mexico_loader.py — ETL CNBV Boletín Estadístico Banca Múltiple → CockroachDB
LatamBanks — country='MX'

Fuente (mensual, Excel en Portafolio de Información):
  https://portafolioinfo.cnbv.gob.mx/PortafolioInformacion/{nombre}

  Nombres habituales (se prueban en orden):
    BE BM {YYYYMM}.xlsx   (recientes, con espacio)
    BE_BM_{YYYYMM}.xlsx
    BE_BM_{YYYYMM}.xlsm
    BE_BM_{YYYYMM}.xls

  Sheet «Pm2» → Principales rubros (Activo, Cartera, Captación, Capital, Resultado neto)

Unidad: millones de MXN → monto_total en pesos enteros (×1_000_000).
ins_cod: hash estable del nombre (+ tabla fija para grandes bancos).

Nota cobertura — Nu México / Nubank:
  El loader solo lee Banca Múltiple (BM). Nu operó como SOFIPO y, hasta el BM
  202605 inclusive, no figura en Pm2. La autorización CNBV como banco (~jul 2026)
  implica que aparecerá cuando el boletín BM lo publique como institución BM;
  no inventamos filas ni pines hasta entonces.

Modos:
  (sin flags)     Incremental
  --month AAAAMM
  --all / --from / --to
  --dry-run
  --wipe
"""
from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import ssl
import time
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import HTTPSHandler, Request, build_opener

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Falta openpyxl. pip install openpyxl") from e

try:
    import xlrd
except ImportError:
    xlrd = None  # solo necesario para .xls legacy

COUNTRY = "MX"
BATCH = 500
BASE = "https://portafolioinfo.cnbv.gob.mx/PortafolioInformacion"
MIN_PERIOD = "201501"
SCALE = 1_000_000  # MDP → MXN

# KPI canónicos desde Pm2
PM2_METRICS = [
    ("Activo total", "TOTAL_ACTIVO", "b1", "Activo total"),
    ("Cartera total", "CARTERA_TOTAL", "b1", "Cartera de crédito total"),
    ("Captación total", "CAPTACION_TOTAL", "b1", "Captación total"),
    ("Capital contable", "CAPITAL_CONTABLE", "b1", "Capital contable"),
    ("Resultado neto", "RESULTADO_NETO", "r1", "Resultado neto"),
]

BANK_CODE_BY_NORM = {
    "bbva mexico": 12,
    "santander": 14,
    "banorte": 72,
    "banamex": 2,
    "hsbc": 21,
    "scotiabank": 44,
    "inbursa": 36,
    "citi mexico": 9,
    "banco del bajio": 30,
    "banco azteca": 127,
    "afirme": 62,
    "monex": 112,
    "banregio": 58,
    "invex": 59,
    "j.p. morgan": 86,
    "jp morgan": 86,
    "banco base": 145,
    "bancoppel": 137,
    "multiva": 132,
    "banca mifel": 42,
    "compartamos": 130,
}

EXCLUDE_NAME_RE = re.compile(r"^\s*sistema\b|total\s+banca", re.I)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("mexico_loader")

_CTX = ssl.create_default_context()
try:
    _CTX.check_hostname = False
    _CTX.verify_mode = ssl.CERT_NONE
except Exception:
    pass
_OPENER = build_opener(HTTPSHandler(context=_CTX))


def fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def ym_to_period(y: int, m: int) -> str:
    return f"{y}{m:02d}"


def iter_periods(start: str, end: str | None = None) -> list[str]:
    from datetime import date

    if end is None:
        today = date.today()
        end = ym_to_period(today.year, today.month)
    y, m = int(start[:4]), int(start[4:6])
    ye, me = int(end[:4]), int(end[4:6])
    out = []
    while (y, m) <= (ye, me):
        out.append(ym_to_period(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def recent_candidate_periods(n_months: int = 4) -> list[str]:
    from datetime import date

    today = date.today()
    y, m = today.year, today.month
    out = []
    for _ in range(n_months):
        out.append(ym_to_period(y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))


def candidate_names(periodo: str) -> list[str]:
    return [
        f"BE BM {periodo}.xlsx",
        f"BE_BM_{periodo}.xlsx",
        f"BE_BM_{periodo}.xlsm",
        f"BE_BM_{periodo}.xls",
        f"BE_BM_ {periodo}.xlsx",
        f"BE_BM _{periodo}.xlsx",
    ]


def file_url(name: str) -> str:
    # quote path segment but keep spaces as %20
    return f"{BASE}/{quote(name)}"


def http_bytes(url: str, retries: int = 3, backoff: float = 3.0) -> bytes:
    last = None
    for attempt in range(1, retries + 1):
        try:
            req = Request(url, headers={"User-Agent": "LatamBanksMX/1.0"})
            with _OPENER.open(req, timeout=180) as r:
                return r.read()
        except (HTTPError, URLError, TimeoutError) as e:
            last = e
            if isinstance(e, HTTPError) and e.code == 404:
                raise
            log.warning("intento %d/%d fallo %s: %s", attempt, retries, url, e)
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise last


def resolve_bulletin(periodo: str) -> tuple[str, bytes]:
    last_err = None
    for name in candidate_names(periodo):
        url = file_url(name)
        try:
            data = http_bytes(url)
            if len(data) < 1000:
                continue
            log.info("dt=%s archivo %s (%d bytes)", periodo, name, len(data))
            return name, data
        except HTTPError as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            continue
    raise FileNotFoundError(f"Sin boletín BM para {periodo}: {last_err}")


def index_exists(periodo: str) -> bool:
    try:
        resolve_bulletin(periodo)
        return True
    except Exception:
        return False


def _list_portafolio_names() -> list[str]:
    """Lista nombres en /PortafolioInformacion vía SharePoint REST (sin auth)."""
    import json

    url = (
        "https://portafolioinfo.cnbv.gob.mx/_api/web/"
        "GetFolderByServerRelativeUrl('/PortafolioInformacion')/Files?$top=5000"
    )
    req = Request(
        url,
        headers={
            "User-Agent": "LatamBanksMX/1.0",
            "Accept": "application/json;odata=verbose",
        },
    )
    with _OPENER.open(req, timeout=180) as r:
        data = json.loads(r.read().decode())
    return [f["Name"] for f in data.get("d", {}).get("results", [])]


def discover_available_periods(start: str = MIN_PERIOD, end: str | None = None) -> list[str]:
    from datetime import date

    if end is None:
        today = date.today()
        end = ym_to_period(today.year, today.month)
    start = max(start, MIN_PERIOD)

    found: set[str] = set()
    try:
        names = _list_portafolio_names()
        for name in names:
            m = re.search(r"BE[\s_]*BM[\s_]*(\d{6})\.(xlsx|xlsm|xls)$", name, re.I)
            if not m:
                continue
            p = m.group(1)
            if start <= p <= end:
                found.add(p)
        log.info("Catálogo MX vía SharePoint: %d nombres, %d en rango", len(names), len(found))
    except Exception as e:
        log.warning("SharePoint list falló (%s); sondeo por mes", e)
        for p in iter_periods(start, end):
            if index_exists(p):
                found.add(p)

    out = sorted(found)
    log.info(
        "Catálogo MX: %d meses (%s .. %s)",
        len(out),
        out[0] if out else "—",
        out[-1] if out else "—",
    )
    return out


def bank_code(name: str) -> int:
    n = fold(name)
    n = re.sub(r"\s*\(antes[^)]*\)\s*", " ", n).strip()
    if n in BANK_CODE_BY_NORM:
        return BANK_CODE_BY_NORM[n]
    for k, v in BANK_CODE_BY_NORM.items():
        if k in n or n in k:
            return v
    h = int(hashlib.md5(n.encode()).hexdigest()[:6], 16)
    return 1000 + (h % 8000)


def _to_int_mdp(v) -> int:
    if v in ("", None, "-"):
        return 0
    try:
        return int(round(float(v) * SCALE))
    except (TypeError, ValueError):
        return 0


def _period_from_date(v) -> str | None:
    if isinstance(v, datetime):
        return f"{v.year}{v.month:02d}"
    if isinstance(v, str) and re.match(r"^\d{4}-\d{2}", v):
        return v[:4] + v[5:7]
    return None


def parse_pm2_openpyxl(data: bytes, periodo: str):
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    if "Pm2" not in wb.sheetnames:
        # algunos boletines usan nombre distinto
        cand = [s for s in wb.sheetnames if fold(s).startswith("pm")]
        if not cand:
            raise RuntimeError(f"Sheet Pm2 ausente: {wb.sheetnames}")
        sh = wb[cand[0]]
    else:
        sh = wb["Pm2"]

    # materializar filas necesarias
    grid = []
    for row in sh.iter_rows(max_col=60, values_only=True):
        grid.append(list(row))
        if len(grid) > 120:
            break

    header_row = None
    date_row = None
    for i, row in enumerate(grid):
        joined = " | ".join(str(c) for c in row if c)
        if "Activo total" in joined and "Capital" in joined:
            header_row = i
            date_row = i + 1
            break
    if header_row is None:
        raise RuntimeError("No se encontró fila de headers Pm2")

    headers = grid[header_row]
    dates = grid[date_row] if date_row < len(grid) else []

    # métrica → columna del período pedido (o la más reciente ≤ período)
    metric_cols: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h or not isinstance(h, str):
            continue
        hf = fold(h)
        matched = None
        for label, cuenta, _tipo, _desc in PM2_METRICS:
            lf = fold(label)
            if hf.startswith(lf) or lf in hf:
                matched = cuenta
                break
        if not matched or matched in metric_cols:
            continue
        # patrón: header en i; valores en i / i+2 / i+4 (con % intercalados)
        best = None
        best_p = None
        for off in (0, 2, 4):
            col = i + off
            if col >= len(dates):
                continue
            p = _period_from_date(dates[col])
            if not p:
                continue
            if p == periodo:
                best = col
                break
            if p <= periodo and (best_p is None or p > best_p):
                best = col
                best_p = p
        if best is None:
            for off in (4, 2, 0):
                col = i + off
                if col < len(dates) and _period_from_date(dates[col]):
                    best = col
                    break
        if best is not None:
            metric_cols[matched] = best

    if "TOTAL_ACTIVO" not in metric_cols or "CAPITAL_CONTABLE" not in metric_cols:
        raise RuntimeError(f"Pm2 incompleto: cols={metric_cols}")

    plan = {cuenta: desc for _l, cuenta, _t, desc in PM2_METRICS}
    tipo_by = {cuenta: tipo for _l, cuenta, tipo, _d in PM2_METRICS}
    rows = []
    inst = []
    seen = set()

    for row in grid[date_row + 1 :]:
        name = row[1] if len(row) > 1 else None
        if not name or not isinstance(name, str):
            continue
        ns = re.sub(r"\s+", " ", name.replace("\n", " ")).strip()
        if not ns or EXCLUDE_NAME_RE.search(ns):
            continue
        # fila de banco: debe tener activo numérico
        act_col = metric_cols["TOTAL_ACTIVO"]
        if act_col >= len(row) or not isinstance(row[act_col], (int, float)):
            continue
        code = bank_code(ns)
        if code in seen:
            code = code + 100
        seen.add(code)
        inst.append((COUNTRY, code, ns))
        for cuenta, col in metric_cols.items():
            val = _to_int_mdp(row[col] if col < len(row) else None)
            rows.append(
                (COUNTRY, periodo, tipo_by[cuenta], code, cuenta, 0, 0, 0, 0, val)
            )

    if not inst:
        raise RuntimeError(f"Sin bancos en Pm2 para {periodo}")
    return inst, rows, plan


def parse_workbook(data: bytes, periodo: str, filename: str = ""):
    lower = filename.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx") and not lower.endswith(".xlsm"):
        if xlrd is None:
            raise SystemExit("Falta xlrd para .xls legacy")
        # Intentar openpyxl primero no sirve; xlrd + reconstrucción limitada
        # Para .xls antiguos sin Pm2 moderno, fallar claro
        book = xlrd.open_workbook(file_contents=data)
        names = book.sheet_names()
        if not any(fold(n).startswith("pm") for n in names):
            raise RuntimeError(f"{periodo}: .xls sin hoja Pm2 ({names[:8]}) — usar xlsx/xlsm")
        # convertir vía valores a grid fake openpyxl-like es complejo; pedir xlsx
        raise RuntimeError(f"{periodo}: preferir .xlsx/.xlsm (archivo .xls legacy)")
    return parse_pm2_openpyxl(data, periodo)


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def get_db_url() -> str:
    url = os.environ.get("COCKROACH_URL")
    if not url:
        raise SystemExit("Falta COCKROACH_URL (.env o entorno).")
    return url


def connect():
    import psycopg2

    return psycopg2.connect(get_db_url())


def upsert(conn, table, cols, updates, conflict, rows):
    import psycopg2.extras

    if not rows:
        return
    cur = conn.cursor()
    sql = (
        f"INSERT INTO {table} ({','.join(cols)}) VALUES %s "
        f"ON CONFLICT ({','.join(conflict)}) DO UPDATE SET "
        + ", ".join(f"{c}=EXCLUDED.{c}" for c in updates)
    )
    for i in range(0, len(rows), BATCH):
        psycopg2.extras.execute_values(cur, sql, rows[i : i + BATCH])
    conn.commit()


def wipe_mx(conn):
    cur = conn.cursor()
    for t in ("datos_financieros", "instituciones", "plan_cuentas", "carga_log"):
        cur.execute(f"DELETE FROM {t} WHERE country=%s", (COUNTRY,))
    conn.commit()
    log.info("México borrado (solo country='MX').")


def get_loaded_periods(conn) -> set[str]:
    cur = conn.cursor()
    cur.execute(
        "SELECT periodo FROM carga_log WHERE country=%s AND estado IN ('ok','alerta_esquema')",
        (COUNTRY,),
    )
    return {r[0] for r in cur.fetchall()}


def load_month(conn, periodo: str) -> None:
    from schema_guard import detect_schema_changes, get_known_accounts, record_schema_result

    name, data = resolve_bulletin(periodo)
    inst_rows, all_rows, plan = parse_workbook(data, periodo, name)
    if not all_rows:
        raise RuntimeError(f"Sin filas de datos para {periodo}")

    known = get_known_accounts(conn, COUNTRY)
    incoming = {r[4] for r in all_rows}
    report = detect_schema_changes(COUNTRY, periodo, incoming, known)

    upsert(
        conn,
        "instituciones",
        ["country", "codigo", "razon_social"],
        ["razon_social"],
        ["country", "codigo"],
        inst_rows,
    )
    upsert(
        conn,
        "plan_cuentas",
        ["country", "cuenta", "descripcion"],
        ["descripcion"],
        ["country", "cuenta"],
        [(COUNTRY, c, d) for c, d in sorted(plan.items())],
    )
    upsert(
        conn,
        "datos_financieros",
        [
            "country",
            "periodo",
            "tipo",
            "ins_cod",
            "cuenta",
            "monto_clp",
            "monto_uf",
            "monto_tc",
            "monto_ext",
            "monto_total",
        ],
        ["monto_total"],
        ["country", "periodo", "tipo", "ins_cod", "cuenta"],
        all_rows,
    )
    upsert(
        conn,
        "carga_log",
        ["country", "periodo", "archivos_procesados", "estado"],
        ["archivos_procesados", "estado"],
        ["country", "periodo"],
        [(COUNTRY, periodo, len(inst_rows), "ok")],
    )
    record_schema_result(conn, COUNTRY, periodo, report)
    log.info(
        "dt=%s OK (%s): %d bancos, %d filas (schema=%s)",
        periodo,
        name,
        len(inst_rows),
        len(all_rows),
        report.get("status"),
    )


def main(argv: Iterable[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="ETL CNBV BM México → CockroachDB")
    ap.add_argument("--wipe", action="store_true")
    ap.add_argument("--month", help="AAAAMM puntual")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--from", dest="from_dt", default=MIN_PERIOD)
    ap.add_argument("--to", dest="to_dt", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(list(argv) if argv is not None else None)

    if args.month:
        periods = [args.month]
    elif args.all or args.to_dt or (args.from_dt and args.from_dt != MIN_PERIOD):
        periods = discover_available_periods(args.from_dt, args.to_dt)
    else:
        periods = [p for p in recent_candidate_periods(4) if index_exists(p)]

    if args.from_dt:
        periods = [p for p in periods if p >= args.from_dt]
    if args.to_dt:
        periods = [p for p in periods if p <= args.to_dt]

    if args.dry_run:
        print(f"Meses objetivo: {len(periods)}")
        for p in periods[:15]:
            try:
                name, _ = resolve_bulletin(p)
                print(f"  {p}  {name}")
            except Exception as e:
                print(f"  {p}  FAIL {e}")
        if periods:
            name, data = resolve_bulletin(periods[-1])
            inst, rows, plan = parse_workbook(data, periods[-1], name)
            print(f"Bancos ({periods[-1]}): {len(inst)}")
            print(" top:", [(c, n) for _, c, n in inst[:8]])
            by = {}
            for r in rows:
                if r[3] == 12 and r[4] in (
                    "TOTAL_ACTIVO",
                    "CARTERA_TOTAL",
                    "CAPTACION_TOTAL",
                    "CAPITAL_CONTABLE",
                    "RESULTADO_NETO",
                ):
                    by[r[4]] = r[9]
            print("BBVA KPIs (MXN):", by)
            print(f"filas={len(rows)} plan={len(plan)}")
        return 0

    conn = connect()
    try:
        if args.wipe:
            wipe_mx(conn)

        if not args.month and not args.all:
            loaded = get_loaded_periods(conn)
            periods = [p for p in periods if p not in loaded]
            if not periods:
                log.info("México al día. Nada nuevo.")
                return 0

        log.info(
            "Cargando %d mes(es): %s%s",
            len(periods),
            periods[:6],
            "…" if len(periods) > 6 else "",
        )
        for p in periods:
            load_month(conn, p)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
